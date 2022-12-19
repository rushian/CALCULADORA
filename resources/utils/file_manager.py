import re
import openpyxl
from openpyxl import load_workbook


# metodo de leitura de celula
def ler_dado_celula(arquivo: str, aba: str, celula: str):
    planilha = load_workbook(arquivo)
    sheet_ranges = planilha[aba]
    return sheet_ranges[celula].value 


def ler_dado_celula_ao_lado(arquivo: str, aba: str, celula: str):

    planilha = load_workbook(arquivo)
    sheet_ranges = planilha[aba]
    coluna = re.search("(\D+)", celula).group()
    linha = int(re.search("(\d+)", celula).group())

    colInt = openpyxl.utils.column_index_from_string(coluna)
    
    return sheet_ranges.cell(row=linha, column=colInt+1).value



def escrever_dado_na_celula(arquivo, aba, celula, valor):
    wkbk = load_workbook(arquivo,read_only=False)
    planilha = wkbk[aba]
    planilha[celula] = valor
    wkbk.save(arquivo)

 
